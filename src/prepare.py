
import openpyxl as excel
import json, math
from copy import deepcopy

TASK_MACHINE  = 0
TASK_DURATION = 1

PRODUCTION_TIME = 'production_time'
PRODUCTION_LINE = 'production_line'
CAPACITY        = 'capacity'
MODEL_TOTALS    = 'total'
MIN_START       = 'min_start'
MAX_END         = 'max_end'
TASK            = 'task'

JOBS        = 'jobs'
HORIZON     = 'horizon'
NORMAL_TIME = 'normal_time'
OVER_TIME   = 'over_time'

ORTOOLS_DATA_FILE = 'data/fab.json'
PROLOG_DATA_FILE  = 'data/fab.pl'
CPLEX_DATA_FILE   = 'data/fab.dat'

LOG = True
WORK_WEEK = 2304

######################### Real Data #########################

def get_data() -> tuple:
    # Data Constants
    FILE = 'data/raw/oi_22_23.xlsx'
    TIMES_SHEET            = 'times'
    TIMES_MODEL_COLUMN     = 1
    TIMES_TIME_COLUMN      = 2
    PRODUCTION_RANGE_SHEET = 'dates'
    RANGE_MODEL_COLUMN     = 1
    RANGE_START_COLUMN     = 2
    RANGE_END_COLUMN       = 3
    LINES_SHEET            = 'lines'
    LINES_LINE_COLUMN      = 1
    LINES_CAPACITY_COLUMN  = 2
    LINES_MODELS_COLUMN    = 3
    TOTAL_SHEET            = 'total'
    TOTAL_MODELS_COLUMN    = 1
    TOTAL_TOTALS_COLUMN    = 3
    
    # Prepare Data
    workbook = excel.load_workbook(FILE)
    models = {}
    lines = {}
    lines_aux = {}

    # Times
    sheet  = workbook[TIMES_SHEET]
    n_rows = sheet.max_row
    for row in range(2, n_rows + 1):
        model      = sheet.cell(row, TIMES_MODEL_COLUMN).value
        time_taken = sheet.cell(row, TIMES_TIME_COLUMN ).value
        models[model] = { PRODUCTION_TIME : time_taken, MODEL_TOTALS: 0, MIN_START: 0, MAX_END: 0}
    
    # Production Range
    sheet = workbook[PRODUCTION_RANGE_SHEET]
    n_rows = sheet.max_row
    for row in range(2, n_rows):
        model     = sheet.cell(row, RANGE_MODEL_COLUMN).value
        min_start = sheet.cell(row, RANGE_START_COLUMN).value * WORK_WEEK
        max_end   = sheet.cell(row, RANGE_END_COLUMN  ).value * WORK_WEEK
        models[model][MIN_START] = min_start
        models[model][MAX_END  ] = max_end
    
    # Lines
    sheet = workbook[LINES_SHEET]
    n_rows = sheet.max_row
    for row in range(2, n_rows + 1):
        line        = sheet.cell(row, LINES_LINE_COLUMN    ).value
        capacity    = sheet.cell(row, LINES_CAPACITY_COLUMN).value
        line_models = sheet.cell(row, LINES_MODELS_COLUMN  ).value
        line_models = line_models.split(', ')
        min_max_models = []
        for line_model in line_models:
            values = line_model.split('-')
            min_max_models.append((int(values[0]), int(values[1])))
        lines[line] = { CAPACITY : capacity }
        lines_aux[line] = min_max_models
    
    # Models
    for model in models:
        models[model][PRODUCTION_LINE] = []
        for line in lines:
            for (min_model, max_model) in lines_aux[line]:
                if min_model <= model <= max_model:
                    models[model][PRODUCTION_LINE].append(line)
    
    # Totals
    sheet = workbook[TOTAL_SHEET]
    n_rows = sheet.max_row
    for row in range(2, n_rows + 1):
        model = sheet.cell(row, TOTAL_MODELS_COLUMN).value
        total = sheet.cell(row, TOTAL_TOTALS_COLUMN).value
        if model in models:
            models[model][MODEL_TOTALS] += total
        else:
            if LOG:
                print(f'Model {model} doesn\'t exist')
    
    # Remove Models that don't have any pieces to produce
    del_models = []
    for model_id, model in models.items():
        if model[MODEL_TOTALS] == 0:
            del_models.append(model_id)
    for del_model in del_models:
        if LOG:
            print(f'Model {del_model} doesn\'t have any pieces to produce')
        del models[del_model]
    
    # Finished
    workbook.close()
    return (models, lines)

def get_jobs(models: dict, lines: dict) -> dict:
    # Each Job has 1 Production Task which may have alternative tasks in different machines

    jobs = {}
    for model_id, model in models.items():
        production_time = model[PRODUCTION_TIME] * model[MODEL_TOTALS]
        alternative_tasks = []
    
        for production_line in model[PRODUCTION_LINE]:
            t = [0, 0]
            t[TASK_MACHINE ] = production_line
            t[TASK_DURATION] = math.ceil(production_time / lines[production_line][CAPACITY])
            alternative_tasks.append(tuple(t))
        
        jobs[model_id] = {TASK: alternative_tasks, MIN_START: model[MIN_START], MAX_END: model[MAX_END]}
    return jobs

######################### Data to Files #########################

def save_data(jobs: dict, n_lines: int, horizon: int = 68000, normal_time: int = 1920, over_time: int = 384,
              ortools_file: str = ORTOOLS_DATA_FILE, prolog_file: str = PROLOG_DATA_FILE,
              cplex_file: str = CPLEX_DATA_FILE) -> None:
    save_ortools(jobs         , horizon, normal_time, over_time, ortools_file)
    save_prolog (jobs, n_lines, horizon, normal_time, over_time,  prolog_file)
    # save_cplex  (jobs, n_lines,   cplex_file)

def save_ortools(jobs: dict, horizon: int, normal_time: int, over_time: int, ortools_file: str) -> None:
    file_content = {JOBS: jobs, HORIZON: horizon, NORMAL_TIME: normal_time, OVER_TIME: over_time}
    with open(ortools_file, 'w') as file:
        json.dump(file_content, file)
    return

def save_prolog(jobs: dict, n_lines: int, horizon: int, normal_time: int, over_time: int, prolog_file: str) -> None:
    lines = get_prolog_lines(deepcopy(jobs), n_lines, horizon, normal_time, over_time)
    
    with open(prolog_file, 'w') as file:
        file.writelines(lines)
    return

def get_prolog_lines(jobs: dict, n_lines: int, horizon: int, normal_time: int, over_time: int) -> list:
    # Each Job has 1 Production Task which may have alternative tasks in different machines

    lines = [
        '% n_machines(-NMachines)\n',
         f':- assertz(n_machines({n_lines})).\n\n',
        '% horizon(-Horizon)\n',
         f':- assertz(horizon({horizon})).\n\n',
        '% normal_time(-NormalTime).\n',
         f':- assertz(normal_time({normal_time})).\n\n',
        '% over_time(-OverTime).\n',
         f':- assertz(over_time({over_time})).\n\n',
        '% job(+JobId, -MinStart, +MaxEnd, +Task) | Task = [AltTask] | AltTask = MachineId-Duration\n'
    ]

    for model_id, info in jobs.items():
        task      = info[TASK]
        min_start = info[MIN_START]
        max_end   = info[MAX_END]

        for alt_task_id, alt_task in enumerate(task):
            task[alt_task_id] = f'{alt_task[TASK_MACHINE]}-{alt_task[TASK_DURATION]}'
        
        line = f':- assertz(job({model_id}-{min_start}-{max_end}-{task})).\n'
        line = line.replace('\'', '')
        lines.append(line)
    return lines

def save_cplex(jobs: dict, n_lines: int, cplex_file: str) -> None:
    lines = get_cplex_lines(jobs, n_lines)
    
    with open(cplex_file, 'w') as file:
        file.writelines(lines)
    return

def get_cplex_lines(jobs: dict, n_lines: int) -> list:
    ops   = [ 'Ops = {   // OperationID, JobID, JobPosition\n'      ]
    modes = [ 'Modes = { // OperationID, Machine, ProcessingTime\n' ]
    
    for job_id, info in jobs.items():
        task      = info[TASK]
        min_start = info[MIN_START]
        max_end   = info[MAX_END]
        
        # TODO: include min_start & max_end
        
        ops.append(f'  <{job_id}, {job_id}, 0>,\n')
        for alt_task in task:
            modes.append(f'  <{job_id}, {alt_task[TASK_MACHINE]}, {alt_task[TASK_DURATION]}>,\n')
    
    ops  .append('};\n\n')
    modes.append('};\n')
    
    lines = [f'Params = <{len(jobs)}, {n_lines}>;\n\n'] # TODO
    lines.extend(ops  )
    lines.extend(modes)
    return lines

######################### Real Data Statistics #########################

def get_duration(job_tasks: list) -> int:
    job_tasks = job_tasks[TASK]
    return list(map(lambda x: x[TASK_DURATION], job_tasks))

def statistics(jobs: dict, models: dict) -> None:
    def n_production_lines(job_tasks):
        production_tasks = job_tasks[1][TASK]
        return len(production_tasks), models[job_tasks[0]][MODEL_TOTALS]

    job_tasks     = list(map(n_production_lines, jobs.items()))
    only_one_line = list(map(lambda x: x[1], filter(lambda job: job[0] == 1, job_tasks)))
    both_lines    = list(map(lambda x: x[1], filter(lambda job: job[0] == 2, job_tasks)))

    print()
    print(f'Number of Different Products: {len(job_tasks)}')
    print(f'Number of Products that can only be produced in 1 line: {len(only_one_line)} - {sum(only_one_line)}')
    print(f'Number of Products that can be produced in both lines:  {len(both_lines   )} - {sum(both_lines   )}')

    min_units = min(min(only_one_line), min(both_lines))
    max_units = max(max(only_one_line), max(both_lines))
    print(f'Units per Product: [{min_units}, {max_units}]')
    
    task_durations = map(get_duration, jobs.values())
    durations = [alt_task_duration for alt_task_durations in task_durations for alt_task_duration in alt_task_durations]
    
    print(f'Product Duration : [{min(durations)}, {max(durations)}]')

def model_statistics(jobs: dict, models: dict, id: int) -> None:
    print()
    print(f'Model {id}')
    print(f'  Production Time per Unit: {models[id][PRODUCTION_TIME]}')
    print(f'  Number of Units:          {models[id][MODEL_TOTALS]}')
    print(f'  Production Lines:         {models[id][PRODUCTION_LINE]}')
    print(f'  Duration:                 {get_duration(jobs[id])}')
    return

if __name__ == '__main__':
    # Get Data
    (models, lines) = get_data()
    jobs = get_jobs(models, lines)
    
    # Save Data
    save_data(jobs, len(lines))
    
    # Show Statistics
    if LOG:
        statistics(jobs, models)
        model_statistics(jobs, models, 128)
        model_statistics(jobs, models, 818)
        print(list(jobs.items())[0])
    
    # Test
    print('Done.')
