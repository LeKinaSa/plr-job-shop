
import openpyxl as excel

def get_data(production_time = 'production_time',
             production_line = 'production_line',
             capacity        = 'capacity'):
    # Data Constants
    FILE = 'data/oi_22_23.xlsx'
    TIMES_SHEET = 'times'
    TIMES_MODEL_COLUMN = 1
    TIMES_TIME_COLUMN  = 2
    LINES_SHEET = 'lines'
    LINES_LINE_COLUMN     = 1
    LINES_CAPACITY_COLUMN = 2
    LINES_MODELS_COLUMN   = 3
    
    # Prepare Data
    workbook = excel.load_workbook(FILE)
    models = {}
    lines = {}
    lines_aux = {}

    # Times
    sheet  = workbook[TIMES_SHEET]
    n_rows = sheet.max_row
    for row in range(2, n_rows + 1):
        model      = sheet.cell(row, TIMES_MODEL_COLUMN).value
        time_taken = sheet.cell(row, TIMES_TIME_COLUMN ).value
        models[model] = { production_time : time_taken}
    
    # Lines
    sheet = workbook[LINES_SHEET]
    n_rows = sheet.max_row
    for row in range(2, n_rows + 1):
        line        = sheet.cell(row, LINES_LINE_COLUMN    ).value
        capacity    = sheet.cell(row, LINES_CAPACITY_COLUMN).value
        line_models = sheet.cell(row, LINES_MODELS_COLUMN  ).value
        line_models = line_models.split(', ')
        min_max_models = []
        for line_model in line_models:
            values = line_model.split('-')
            min_max_models.append((int(values[0]), int(values[1])))
        lines[line] = { capacity : capacity }
        lines_aux[line] = min_max_models
    
    # Models
    for model in models:
        models[model][production_line] = []
        for line in lines:
            for (min_model, max_model) in lines_aux[line]:
                if min_model <= model <= max_model:
                    models[model][production_line].append(line)
    
    # Finished
    workbook.close()
    return (models, lines)

if __name__ == '__main__':
    data = get_data()
    print(data)
